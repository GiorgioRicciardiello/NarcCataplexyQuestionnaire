from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from config.config import config
from openpyxl.styles import Font, Alignment, Border, Side


def format_csv_to_excel(path_csv: Path):
    # Load CSV
    df = pd.read_csv(path_csv)

    # Set output path
    path_excel = path_csv.with_suffix('.xlsx')
    df.to_excel(path_excel, index=False)

    # Load workbook
    wb = load_workbook(path_excel)
    ws = wb.active

    # Style definitions
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_side = Side(border_style="thin", color="000000")
    header_border = Border(top=thin_side, bottom=thin_side)
    bottom_border = Border(bottom=thin_side)

    # Dimensions
    max_row = ws.max_row
    max_col = ws.max_column

    # Style header
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = header_border

    # Style body
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = center_align
            if col == 1:
                cell.font = bold_font

    # Bottom border on last row
    for col in range(1, max_col + 1):
        ws.cell(row=max_row, column=col).border = bottom_border

    # Auto-fit column width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Save final version
    wb.save(path_excel)
    return path_excel

if __name__ == '__main__':
    # Get all CSV files in the folder
    folder = config.get('root_path').joinpath('src')
    csv_files = list(folder.glob("*.csv"))
    # edit the csvs into the excel format we like
    for path_ in csv_files:
        format_csv_to_excel(path_csv=path_)